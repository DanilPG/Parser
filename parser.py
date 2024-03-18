from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
import urllib3 # импортируем библиотеку urllib3


def parse():
    urllib3.disable_warnings()  # отключаем варнинги чтобы пустило на сайт
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36'
    }
    url = 'https://omsk.rabota.ru/vacancy/?query=python&sort=relevance'  # ссылка
    page = requests.get(url, headers=headers, verify=False)  # запрос html
    print(page.status_code)
    soup = BeautifulSoup(page.text, "html.parser")
    title = soup.findAll('h3', class_='vacancy-preview-card__title')
    link = soup.findAll('a', class_='vacancy-preview-card__title_border', target='_blank')
    wb = Workbook()
    ws = wb.active
    i=1
    for data in title:
        a=data.text
        ws.cell(row=i, column=1, value=a)
        i+=1
    j=1
    for data2 in link:
        href='https://omsk.rabota.ru/'+data2.get('href')
        ws.cell(row=j, column=9, value=href)
        j += 1
    wb.save('vacancies.xlsx')
    print("Файл успешно создан")